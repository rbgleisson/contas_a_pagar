# core/export_excel.py — Exportação Excel com data BR + Relatório Mensal por Categoria
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, numbers

from . import models

def _formatar_data_br(data_str: str) -> str:
    try:
        return datetime.strptime(data_str, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return data_str or ""

def _autoajustar_colunas(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

def _preencher_sheet(ws, linhas, tipo: str):
    """
    linhas: lista de dicts (pagar/receber) com chaves:
      descricao, valor, vencimento(YYYY-MM-DD), conta_nome, categoria, pago/recebido
    tipo: "pagar" ou "receber" (define o texto do status)
    """
    ws.append(["Descrição", "Valor", "Data", "Conta", "Categoria", "Status"])
    header = ws[1]
    for c in header:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for it in linhas:
        status_txt = (
            "Pago" if (tipo == "pagar" and it.get("pago")) else
            "Recebido" if (tipo == "receber" and it.get("recebido")) else
            "Pendente"
        )
        data_br = _formatar_data_br(it.get("vencimento", ""))
        ws.append([
            it.get("descricao", ""),
            float(it.get("valor", 0.0)),
            data_br,
            it.get("conta_nome", ""),
            it.get("categoria", ""),
            status_txt
        ])

    # Formatar coluna Valor como moeda (R$)
    # Coluna 2: Valor, Coluna 3: Data (texto BR)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = u'R$ #,##0.00'

    _autoajustar_colunas(ws)

def export_to_excel(contas_a_pagar: list, contas_a_receber: list):
    """
    Exporta as listas já carregadas da GUI (mantido por compatibilidade).
    Datas saem em BR.
    """
    try:
        wb = Workbook()
        ws_pg = wb.active
        ws_pg.title = "Pagar"
        _preencher_sheet(ws_pg, contas_a_pagar, "pagar")

        ws_rc = wb.create_sheet("Receber")
        _preencher_sheet(ws_rc, contas_a_receber, "receber")

        out = Path.cwd() / "export_financeiro.xlsx"
        wb.save(out)
        return True, f"Arquivo gerado: {out}"
    except Exception as e:
        return False, f"Falha ao exportar: {e}"

def export_monthly_report(mes: int, ano: int, categoria: str | None = None):
    """
    Gera um relatório mensal (mês/ano) filtrado por categoria (ou todas) em Excel.
    Busca direto do banco via models.search_pagar/search_receber.
    """
    try:
        cat = None if (not categoria or categoria.lower() == "todas") else categoria

        pagar = models.search_pagar(mes=mes, ano=ano, categoria=cat)
        receber = models.search_receber(mes=mes, ano=ano, categoria=cat)

        wb = Workbook()
        ws_pg = wb.active
        ws_pg.title = "Pagar"
        _preencher_sheet(ws_pg, pagar, "pagar")

        ws_rc = wb.create_sheet("Receber")
        _preencher_sheet(ws_rc, receber, "receber")

        cat_slug = "Todas" if cat is None else cat.replace(" ", "_")
        out_name = f"Relatorio_{ano}-{int(mes):02d}_{cat_slug}.xlsx"
        out = Path.cwd() / out_name
        wb.save(out)
        return True, f"Relatório gerado: {out}"
    except Exception as e:
        return False, f"Falha ao gerar relatório: {e}"
